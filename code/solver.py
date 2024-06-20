import datetime
import os
import time
import matplotlib
import torch
import torch.nn as nn
import torch.optim as optim
import numpy as np
import matplotlib.pyplot as plt
from metric import compute_measure
import openpyxl
from whaformer import WHAformer as whaformer
from losses import CharbonnierLoss
from fvcore.nn import FlopCountAnalysis, parameter_count_table
from torch.optim.lr_scheduler import CosineAnnealingLR




def split_arr(arr,patch_size,stride=32):
    pad = (16, 16, 16, 16)
    arr = nn.functional.pad(arr, pad, "constant", 0)
    _,_,h,w = arr.shape
    num = h//stride - 1
    arrs = torch.zeros(num*num,1,patch_size,patch_size)

    for i in range(num):
        for j in range(num):
            arrs[i*num+j,0] = arr[0,0,i*stride:i*stride+patch_size,j*stride:j*stride+patch_size]
    return arrs

def agg_arr(arrs, size, stride=32):
    arr = torch.zeros(size, size)
    n,_,h,w = arrs.shape
    num = size//stride
    for i in range(num):
        for j in range(num):
            arr[i*stride:(i+1)*stride,j*stride:(j+1)*stride] = arrs[i*num+j,:,16:48,16:48]
  #return arr
    return arr.unsqueeze(0).unsqueeze(1)



class Solver_whaformer(object):
    def __init__(self, args, dataloader, dataloader_test):

        self.model_name = 'whaformer'

        # set_model
        self.init_lr = args.init_lr

        # train
        self.num_epochs = args.num_epochs
        self.dataloader = dataloader
        self.dataloader_test = dataloader_test
        self.print_iters = args.print_iters
        self.decay_iters = args.decay_iters
        self.multi_gpu = args.multi_gpu

        # test
        self.result_fig = args.result_fig
        self.save_path = args.save_path
        self.test_epoch = args.test_epoch
        self.save_fig_iters = args.save_fig_iters
        self.patch_n = args.patch_n
        self.patch_size = args.patch_size
        self.norm_range_min = args.norm_range_min
        self.norm_range_max = args.norm_range_max
        self.trunc_min = args.trunc_min
        self.trunc_max = args.trunc_max
        self.val_epoch = args.val_epoch
        self.dim = args.dim

        if args.device:
            self.device = torch.device(args.device)
        else:
            self.device = torch.device('cuda:0' if torch.cuda.is_available() else 'cpu')

        self.set_model()

        self.current_epoch = 0
        self.best_epoch = 0
        self.best_score = 0


    def set_model(self):
        former = whaformer(img_size=64, embed_dim=self.dim, depths=[2, 2, 2, 2, 2, 2, 2], win_size=8, mlp_ratio=4., token_projection='linear', token_mlp='mlp', modulator=True, shift_flag=False)
        u_optimizer = torch.optim.Adam(former.parameters(), self.init_lr, weight_decay=1e-4)
        self.metric = CharbonnierLoss()
        self.whaformer = former
        self.whaformer = nn.DataParallel(self.whaformer, device_ids=[0, 1])

        self.whaformer = former.to(self.device)
        self.u_optimizer = u_optimizer
        self.lr_scheduler = CosineAnnealingLR(self.u_optimizer, T_max=self.num_epochs, eta_min=1e-6)

    def lr_decay(self):
        lr = self.init_lr * 0.5
        for param_group in self.u_optimizer.param_groups:
            param_group['lr'] = lr


    def save_model(self, epoch_, loss_=None):
        f = os.path.join(self.save_path, 'WHAFormer_{}epoch.ckpt'.format(epoch_))
        torch.save(self.whaformer.state_dict(), f)
        if loss_:
            f_loss = os.path.join(self.save_path, 'WHAFormer_loss_{}epoch.npy'.format(epoch_))
            np.save(f_loss, np.array(loss_))

    def load_model(self, epoch_):
        f = os.path.join(self.save_path, 'WHAFormer_{}epoch.ckpt'.format(epoch_))
        self.whaformer.load_state_dict(torch.load(f))

    def save_fig(self, x, y, pred, fig_name, original_result, pred_result):

        x, y, pred = x.numpy(), y.numpy(), pred.numpy()

        f, ax = plt.subplots(1, 3, figsize=(30, 10))
        ax[0].imshow(x, cmap=plt.gray(), vmin=self.trunc_min, vmax=self.trunc_max)
        ax[0].set_title('Low-dose', fontsize=30)
        ax[0].set_xlabel("PSNR: {:.4f}\nSSIM: {:.4f}\nRMSE: {:.4f}".format(original_result[0],
                                                                           original_result[1],
                                                                           original_result[2]), fontsize=20)
        ax[1].imshow(pred, cmap=plt.gray(), vmin=self.trunc_min, vmax=self.trunc_max)
        ax[1].set_title('Result', fontsize=30)
        ax[1].set_xlabel("PSNR: {:.4f}\nSSIM: {:.4f}\nRMSE: {:.4f}".format(pred_result[0],
                                                                           pred_result[1],
                                                                           pred_result[2]), fontsize=20)
        ax[2].imshow(y, cmap=plt.gray(), vmin=self.trunc_min, vmax=self.trunc_max)
        ax[2].set_title('Full-dose', fontsize=30)

        f.savefig(os.path.join(self.save_path, 'fig', 'result_{}.png'.format(fig_name)))
        plt.close()

    def save_generated_figure(self, x, y, pred, fig_name, original_result, pred_result):
        x, y, pred = x.numpy(), y.numpy(), pred.numpy()
        plt.axis('off')
        plt.imshow(pred, cmap=plt.gray(), vmin=self.trunc_min, vmax=self.trunc_max)
        plt.savefig(os.path.join(self.save_path, 'fig', 'epoch_of{}_result_{}_generated.png'.format(self.current_epoch,fig_name)), bbox_inches='tight', pad_inches=-0.1)

        plt.close()


    def denormalize_(self, image):
        image = image * (self.norm_range_max - self.norm_range_min) + self.norm_range_min
        return image

    def trunc(self, mat):
        mat[mat <= self.trunc_min] = self.trunc_min
        mat[mat >= self.trunc_max] = self.trunc_max
        return mat

    def train(self):

        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet.append(['EPOCH', 'PSNR', 'SSIM', 'RMSE', self.save_path, datetime.datetime.now()])
        wb.save(self.model_name + '.xlsx')

        self.whaformer.train()

        train_losses = []
        total_iters = 0
        excel_index = 2
        start_time = time.time()



        for epoch in range(1, self.num_epochs):

            self.current_epoch += 1

            for iter_, (x, y) in enumerate(self.dataloader):
                total_iters += 1
                # add 1 channel
                x = x.unsqueeze(0).float().to(self.device)
                y = y.unsqueeze(0).float().to(self.device)

                if self.patch_size:
                    x = x.view(-1, 1, self.patch_size, self.patch_size)
                    y = y.view(-1, 1, self.patch_size, self.patch_size)


                self.u_optimizer.zero_grad()
                self.whaformer.zero_grad()
                pred = self.whaformer(x)
                loss = self.metric(pred, y)

                loss.backward()
                self.u_optimizer.step()

                train_losses.append(loss.item())

                # print
                if total_iters % self.print_iters == 0:
                    print(
                        "STEP [{}], EPOCH [{}/{}], ITER [{}/{}], TIME [{:.1f}s]\nLOSS: {:.8f}".format(
                            total_iters, epoch, self.num_epochs, iter_ + 1, len(self.dataloader),
                                                                 time.time() - start_time, loss.item()))
                # learning rate decay
                if total_iters % self.decay_iters == 0:
                    self.lr_decay()

            # self.lr_scheduler.step()

            # vali
            if epoch > self.val_epoch:
                self.val(epoch)
            # save model
            if epoch > self.val_epoch:
                self.save_model(epoch, train_losses)




        finish_time = time.time() - start_time
        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        # sheet['F1'].value = finish_time
        sheet.append(['finish_time：', finish_time])
        wb.save(self.model_name + '.xlsx')

    def val(self, test_epoch):
        # compute PSNR, SSIM, RMSE
        pred_psnr_avg, pred_ssim_avg, pred_rmse_avg = 0, 0, 0

        total_iters = 0

        self.whaformer.eval()

        with torch.no_grad():
            for i, (x, y) in enumerate(self.dataloader_test):

                total_iters += 1

                shape_ = x.shape[-1]
                x = x.unsqueeze(0).float().to(self.device)
                y = y.unsqueeze(0).float().to(self.device)

                arrs = split_arr(x, 64).to(self.device)  ## split to image patches for test into 4 patches
                arrs[0:64] = self.whaformer(arrs[0:64])
                arrs[64:2 * 64] = self.whaformer(arrs[64:2 * 64])
                arrs[2 * 64:3 * 64] = self.whaformer(arrs[2 * 64:3 * 64])
                arrs[3 * 64:4 * 64] = self.whaformer(arrs[3 * 64:4 * 64])
                pred = agg_arr(arrs, 512).to(self.device)

                # denormalize, truncate
                x = self.trunc(self.denormalize_(x.view(shape_, shape_).cpu().detach()))
                y = self.trunc(self.denormalize_(y.view(shape_, shape_).cpu().detach()))
                pred = self.trunc(self.denormalize_(pred.view(shape_, shape_).cpu().detach()))

                data_range = self.trunc_max - self.trunc_min

                original_result, pred_result = compute_measure(x, y, pred, data_range)
                pred_psnr_avg += pred_result[0]
                pred_ssim_avg += pred_result[1]
                pred_rmse_avg += pred_result[2]

                if total_iters % self.save_fig_iters == 0:
                    if self.result_fig:
                        self.save_generated_figure(x, y, pred, i, original_result, pred_result)

            print('\n')
            pred_psnr_avg_final = pred_psnr_avg / len(self.dataloader_test)
            pred_ssim_avg_final = pred_ssim_avg / len(self.dataloader_test)
            pred_rmse_avg_final = pred_rmse_avg / len(self.dataloader_test)

            if( pred_psnr_avg_final > self.best_score):
                self.best_score = pred_psnr_avg_final
                self.best_epoch = self.current_epoch

            print('bset_epoch {} bset_score {}\n'.format(self.best_epoch, self.best_score))

            print('After learning\nPSNR avg: {:.4f} \nSSIM avg: {:.4f} \nRMSE avg: {:.4f}\n'.format(
                pred_psnr_avg_final, pred_ssim_avg_final,
                pred_rmse_avg_final))
            wb = openpyxl.load_workbook(self.model_name + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.append([test_epoch, pred_psnr_avg_final, pred_ssim_avg_final, pred_rmse_avg_final])
            wb.save(self.model_name + '.xlsx')

    def test(self):
        wb = openpyxl.load_workbook(self.model_name + '.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')
        sheet.append(['test'])
        sheet.append(['PSNR', 'SSIM', 'RMSE'])
        wb.save(self.model_name + '.xlsx')
        del self.whaformer

        # load
        former = whaformer(img_size=64, embed_dim=self.dim, depths=[2, 2, 2, 2, 2, 2, 2], win_size=8, mlp_ratio=4., token_projection='linear', token_mlp='mlp', modulator=True, shift_flag=False)

        self.whaformer = former.to(self.device)
        self.load_model(self.test_epoch)


        tensor = (torch.rand(1, 1, 64, 64).to(self.device),)

        flops = FlopCountAnalysis(self.whaformer, tensor)
        print("FLOPs: ", flops.total())

        print('parameters of generator: \n', parameter_count_table(self.whaformer))

        # compute PSNR, SSIM, RMSE
        ori_psnr_avg, ori_ssim_avg, ori_rmse_avg = 0, 0, 0
        pred_psnr_avg, pred_ssim_avg, pred_rmse_avg = 0, 0, 0

        total_iters = 0
        start_time = 0
        time_flag = 0

        self.whaformer.eval()

        with torch.no_grad():
            for i, (x, y) in enumerate(self.dataloader_test):

                if time_flag == 0:
                    start_time = time.time()
                    time_flag = 1

                total_iters += 1
                shape_ = x.shape[-1]
                x = x.unsqueeze(0).float().to(self.device)
                y = y.unsqueeze(0).float().to(self.device)

                arrs = split_arr(x, 64).to(self.device)  ## split to image patches for test into 4 patches
                arrs[0:64] = self.whaformer(arrs[0:64])
                arrs[64:2 * 64] = self.whaformer(arrs[64:2 * 64])
                arrs[2 * 64:3 * 64] = self.whaformer(arrs[2 * 64:3 * 64])
                arrs[3 * 64:4 * 64] = self.whaformer(arrs[3 * 64:4 * 64])
                pred = agg_arr(arrs, 512).to(self.device)

                # denormalize, truncate
                x = self.trunc(self.denormalize_(x.view(shape_, shape_).cpu().detach()))
                y = self.trunc(self.denormalize_(y.view(shape_, shape_).cpu().detach()))
                pred = self.trunc(self.denormalize_(pred.view(shape_, shape_).cpu().detach()))

                data_range = self.trunc_max - self.trunc_min

                original_result, pred_result = compute_measure(x, y, pred, data_range)
                ori_psnr_avg += original_result[0]
                ori_ssim_avg += original_result[1]
                ori_rmse_avg += original_result[2]
                pred_psnr_avg += pred_result[0]
                pred_ssim_avg += pred_result[1]
                pred_rmse_avg += pred_result[2]

                # save result figure
                if total_iters % self.save_fig_iters == 0:
                    if self.result_fig:
                        self.save_generated_figure(x, y, pred, i, original_result, pred_result)

            print('\n')

            pred_psnr_avg_final = pred_psnr_avg / len(self.dataloader_test)
            pred_ssim_avg_final = pred_ssim_avg / len(self.dataloader_test)
            pred_rmse_avg_final = pred_rmse_avg / len(self.dataloader_test)

        
            total_finish_time = time.time() - start_time
            finish_time = total_finish_time / len(self.dataloader_test)
            print('total running time: ', total_finish_time, '\n average_running time:', finish_time)

            print('Low dose\nPSNR avg: {:.4f} \nSSIM avg: {:.4f} \nRMSE avg: {:.4f}'.format(
                ori_psnr_avg / len(self.dataloader_test), ori_ssim_avg / len(self.dataloader_test),
                ori_rmse_avg / len(self.dataloader_test)))
            print('After learning\nPSNR avg: {:.4f} \nSSIM avg: {:.4f} \nRMSE avg: {:.4f}\n'.format(
                pred_psnr_avg_final, pred_ssim_avg_final,
                pred_rmse_avg_final))

            wb = openpyxl.load_workbook(self.model_name + '.xlsx')
            sheet = wb.get_sheet_by_name('Sheet1')
            sheet.append([pred_psnr_avg_final, pred_ssim_avg_final, pred_rmse_avg_final])
            wb.save(self.model_name + '.xlsx')